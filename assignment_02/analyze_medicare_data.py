# -*- coding: utf-8 -*-
"""
Created on Tue July 07 19:00:00 2017

@author: Gaurav
"""

import os 
import glob 
import requests
import openpyxl
import sqlite3 

#url for CSV FLAT zip file from data.medicare.gov for Hospital Compare data
url_hospital =  "https://data.medicare.gov/views/bg9k-emty/files/0a9879e0-3312-4719-a1db-39fd114890f1?content_type=application%2Fzip%3B%20charset%3Dbinary&filename=Hospital_Revised_Flatfiles.zip"

#create a directory 
staging_dir_name = "staging"
fix_staging_dir_name = "staging_fixed"  #csv files stored in this directory after null fixup
lstState = [] #the list variable to store state list
lstMeasure = [] #Store measure id

def createDirectory(dir_name):
    if not os.path.isdir(dir_name): #directory exists or not check
        #create a directory as it does not exists
        os.mkdir(dir_name)
        
        
#This function is used to access the dataset URL, fetch the data and load it into the directory
def fetchExtractOnlineDataset():
    import zipfile
    # using os.path.join will make your code OS independent
    zip_file_name = os.path.join(staging_dir_name,"hospital_compare_data.zip")
    if not os.path.isfile(zip_file_name):
        r = requests.get(url_hospital)
        zf = open(zip_file_name, "wb")
        zf.write(r.content) # write content to hospital_compare_data.zip file
        zf.close()
    z = zipfile.ZipFile(zip_file_name, 'r') #read the zip file
    z.extractall(staging_dir_name) #extract the zip file
    z.close()

#read every csv file and write a new file which exclude null from original file
def removeNullFromFile():
    glob_dir = os.path.join(staging_dir_name, "*.csv")
    for file_name in glob.glob(glob_dir):
        fn =  os.path.join(staging_dir_name,os.path.basename(file_name))
        in_fp = open(fn, 'rt', encoding ='cp1252')   #rt = read text
        input_data = in_fp.read()
        in_fp.close()
        #getting file name 
        new_file_name = os.path.basename(file_name)
        createDirectory(fix_staging_dir_name)
        #store files into new directory after null fixup
        ofn =  os.path.join(fix_staging_dir_name,new_file_name)
        out_fp = open(ofn, 'wt', encoding ='utf-8')  # wt = write text
        for c in input_data:
            if c != '\0':
                out_fp.write(c)
        out_fp.close()

#this function is used to transform given value into the required format
def transformName(input_name, name_type):
    new_input_name = input_name
    #Convert all letters to lower case
    new_input_name = new_input_name.lower()
    #Replace each blank “ “ with an underscore “_”
    new_input_name = new_input_name.strip()
    new_input_name = new_input_name.replace(" ", "_")
    #Replace each dash or hyphen “-“ with an underscore “_”
    new_input_name = new_input_name.replace("-", "_")
    #Replace each percent sign “%” with the string “pct”
    new_input_name = new_input_name.replace("%", "pct")
    #Replace each forward slash “/” with an underscore “_”
    new_input_name = new_input_name.replace("/", "_")
    #If a table name starts with anything other than a letter “a” through “z” then 
    # prepend “t_” to the front of the table name
    if name_type == "tableName":
        if not new_input_name[0].isalpha():
            new_input_name = "t_" + new_input_name
    else:
        if not new_input_name[0].isalpha():
            # its column so prepend "c_"
            new_input_name = "c_" + new_input_name
    return new_input_name

#This function will create database and table definition
def createDatabaseAndTable():
    try:
        import csv
        conn = sqlite3.connect("medicare_hospital_compare.db")
        c1 = conn.cursor()
        glob_dir = os.path.join(fix_staging_dir_name, "*.csv")
        for file_name in glob.glob(glob_dir):
            #This file is corrupt hence ignored
            if not os.path.basename(file_name) == "FY2015_Percent_Change_in_Medicare_Payments.csv":
                new_table_name = transformName(os.path.splitext(os.path.basename(file_name))[0], "tableName")
                file_path = os.path.join(fix_staging_dir_name,os.path.basename(file_name))
                with open(file_path, "r", encoding ='utf-8') as f:
                    reader = csv.reader(f)
                    col_list = next(reader)
                #prepare table creation query
                sql_str_create = """create table if not exists """ + new_table_name+ """ ( """
                for col_name in col_list:
                    sql_str_create += transformName(col_name, "column") + " text, "
                #remove last extra delimiter
                sql_str_create = sql_str_create[:-2]
                sql_str_create += " ) "
                sql_str_drop = """drop table if exists """ + new_table_name 
                c1.execute(sql_str_drop)
                c1.execute(sql_str_create)
                # table insert query 
                sql_str_insert = """insert into  """ + new_table_name+ """ ( """ + ', '.join( transformName(row, 'column') for row in col_list)+ """ ) values ( """+ ''.join('?, ' for col in col_list)
                sql_str_insert = sql_str_insert[:-2] + """ ) """
                #print(new_table_name)
                #print(sql_str_insert)
                #insert rows into table
                with open(file_path, "r", encoding ='utf-8') as f:
                    reader = csv.reader(f)
                    next(f)
                    for row in reader:                       
                        if len(row) <= len(col_list):
                            #empty rows binding issue resolution
                            if len(row) == 1 :
                                #excluding empty rows
                                if row[0].strip() != '':
                                    counter = len(col_list) - len(row)
                                    while(counter > 0):
                                        row.append('')
                                        counter = counter - 1
                                    #print(new_table_name)
                                    #print(row)  
                                    sql_row = tuple(row)
                                    c1.execute(sql_str_insert, sql_row)
                            else:
                                #rows and column are equal
                                counter = len(col_list) - len(row)
                                while(counter > 0):
                                    row.append('')
                                    counter = counter -1
                                #print(new_table_name)
                                #print(row)  
                                sql_row = tuple(row)
                                c1.execute(sql_str_insert, sql_row)
                        
                conn.commit()
        conn.close()
    finally:
        conn.close() #close database connection
        

# This function will create excel workbooks
def createInHouseHospitalRankingWorkbook():
    #MS Excel Workbook of In House Proprietary Hospital Rankings and Focus List of States
    k_url = 'http://kevincrook.com/utd/hospital_ranking_focus_states.xlsx'
    r = requests.get(k_url)
    xf = open("hospital_ranking_foucus_states.xlsx", "wb") 
    xf.write(r.content)
    xf.close()

# This function will create nationwide adn statewise ranking excel file.
def createHospitalRankingWorkbook():
    wb = openpyxl.Workbook()
    sheet_1 = wb.create_sheet("Nationwide")
    #add headers to excel sheet
    sheet_1.cell(row = 1 , column =1, value = "Provider ID")
    sheet_1.cell(row = 1 , column =2, value = "Hospital Name")
    sheet_1.cell(row = 1 , column =3, value = "City")
    sheet_1.cell(row = 1 , column =4, value = "State")
    sheet_1.cell(row = 1 , column =5, value = "County")
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    #load data for top 100 hospital
    wb2 = openpyxl.load_workbook("hospital_ranking_foucus_states.xlsx")
    ranking_sheet = wb2.get_sheet_by_name("Hospital National Ranking")
    focusState_sheet = wb2.get_sheet_by_name("Focus States")
    list_provider_id = []
    i= 2 #start with 2 to ignore header
    while ranking_sheet.cell(row = i, column = 2).value <= 100 :
        list_provider_id.append(ranking_sheet.cell(row = i, column = 1).value)
        i += 1
        #write data into hospital_ranking file
    try:
        #connect to database
        conn = sqlite3.connect("medicare_hospital_compare.db")
        c1 = conn.cursor()
        sql_str_create = """create table if not exists hospital_ranking_foucus_states (provider_id text, ranking integer )"""
        sql_str_drop = """drop table if exists hospital_ranking_foucus_states """  
        c1.execute(sql_str_drop)
        c1.execute(sql_str_create)
        i= 2 #start with 2 to ignore header
        sql_str_insert = """ insert into hospital_ranking_foucus_states values (?,?)"""
        for rowidx in range(2,ranking_sheet.max_row + 1):
            #insert ranking and provider id into table for temporary purpose.
            c1.execute(sql_str_insert, (ranking_sheet.cell(row = rowidx, column = 1).value, ranking_sheet.cell(row = rowidx, column = 2).value))
        conn.commit()
        """#sql_select_str = "select provider_id, hospital_name, city, state, county_name from hospital_general_information where provider_id in ({})".format(','.join('?' * len(list_provider_id)))
        sql_tupple = list(list_provider_id) 
        rows = c1.execute(sql_select_str, sql_tupple) """
        sql_select_str = """select hr.provider_id, hg.hospital_name, hg.city, hg.state, hg.county_name
        from hospital_ranking_foucus_states hr join hospital_general_information hg 
        on hr.provider_id = hg.provider_id  order by hr.ranking limit 100;"""
        rows = c1.execute(sql_select_str)
        #push Nationawide data into excel file
        for row in rows:
            ws = wb.active 
            ws.append(row)
        #prepare statewise data
        global lstState
        for i in range(2,focusState_sheet.max_row + 1): #max_row + 1 gives all rows
            lstState.append([focusState_sheet.cell(row = i, column = 1).value, focusState_sheet.cell(row = i, column = 2).value])
        import operator
        lstState.sort(key=operator.itemgetter(0))
        for item in lstState:
            sheet_2 = wb.create_sheet(item[0])
            #add headers to excel sheet
            sheet_2.cell(row = 1 , column =1, value = "Provider ID")
            sheet_2.cell(row = 1 , column =2, value = "Hospital Name")
            sheet_2.cell(row = 1 , column =3, value = "City")
            sheet_2.cell(row = 1 , column =4, value = "State")
            sheet_2.cell(row = 1 , column =5, value = "County")
            # the below query will fetch state wise top 100 rows based on rank
            sql_select_str = """select hr.provider_id, hg.hospital_name, hg.city, hg.state, hg.county_name
            from hospital_ranking_foucus_states hr join hospital_general_information hg 
            on hr.provider_id = hg.provider_id where hg.state = '"""+ str(item[1])+ """' order by hr.ranking limit 100;"""
            rows = c1.execute(sql_select_str)
            state_sheet = wb.get_sheet_by_name(item[0])
            #push statewise data into excel file
            for row in rows:                
                state_sheet.append(row)
        sql_str_drop = """drop table if exists hospital_ranking_foucus_states """  
        c1.execute(sql_str_drop)
    finally:
        conn.close()
    wb.save("hospital_ranking.xlsx")
    wb.close()         


# This function will create measures statistics.
def measureStatistics():
    wb = openpyxl.Workbook()
    sheet_1 = wb.create_sheet("Nationwide")
    #add headers to excel sheet
    sheet_1.cell(row = 1 , column =1, value = "Measure ID")
    sheet_1.cell(row = 1 , column =2, value = "Measure Name")
    sheet_1.cell(row = 1 , column =3, value = "Minimum")
    sheet_1.cell(row = 1 , column =4, value = "Maximum")
    sheet_1.cell(row = 1 , column =5, value = "Average")
    sheet_1.cell(row = 1 , column =6, value = "Standard Deviation")
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    try:
        #connect to database
        conn = sqlite3.connect("medicare_hospital_compare.db")
        c1 = conn.cursor()
        #get list of measures
        sql_select_str = """select measure_id, measure_name, min(score), max(score), avg(score) from  timely_and_effective_care___hospital 
        where CAST(score as integer) <> 0 group by measure_id, measure_name order by measure_id"""
        rows = c1.execute(sql_select_str)
        #push statewise data into excel file
        global lstMeasure
        for row in rows:                
            ws = wb.active
            lstMeasure.append(row[0])
            ws.append(row)     
        global lstState
        for item in lstState:
            sheet_2 = wb.create_sheet(item[0])
            #add headers to excel sheet
            sheet_2.cell(row = 1 , column =1, value = "Measure ID")
            sheet_2.cell(row = 1 , column =2, value = "Measure Name")
            sheet_2.cell(row = 1 , column =3, value = "Minimum")
            sheet_2.cell(row = 1 , column =4, value = "Maximum")
            sheet_2.cell(row = 1 , column =5, value = "Average")
            sheet_2.cell(row = 1 , column =6, value = "Standard Deviation")
            # the below query will fetch state wise records
            sql_select_str = """select measure_id, measure_name, min(score), max(score), avg(score) from  timely_and_effective_care___hospital 
            where CAST(score as integer) <> 0 and state = '"""+ str(item[1]) + """' group by measure_id, measure_name order by measure_id"""
            rows = c1.execute(sql_select_str)
            state_sheet = wb.get_sheet_by_name(item[0])
            #push statewise data into excel file
            for row in rows:                
                state_sheet.append(row)
    finally:
        conn.close()
    wb.save("measures_statistics.xlsx")
    wb.close() 

#calculate standard deviation
def calculateStdDev():
    try:
        import statistics
        #connect to database
        conn = sqlite3.connect("medicare_hospital_compare.db")
        c1 = conn.cursor()
        wb2 = openpyxl.load_workbook("measures_statistics.xlsx")
        Nationwide_sheet = wb2.get_sheet_by_name("Nationwide")
        for rowidx in range(2,Nationwide_sheet.max_row + 1):
            #get measure id value amd fetch measure specific data 
            item = Nationwide_sheet.cell(row = rowidx, column = 1).value
            sql_select_main_str = """select  score from  timely_and_effective_care___hospital 
            where CAST(score as integer) <> 0 and measure_id = '"""+ item + """'"""
            select_rows = c1.execute(sql_select_main_str)
            Nationwide_sheet.cell(row = rowidx, column = 6).value = statistics.stdev(int(i[0]) for i in select_rows)
        global lstState
        #loop through state sheets
        for item in lstState:
            state_sheet = wb2.get_sheet_by_name(item[0])
            for rowidx in range(2,state_sheet.max_row + 1):
                measureid = state_sheet.cell(row = rowidx, column = 1).value
                # fetch state and measure specific data
                sql_select_main_str = """select  score from  timely_and_effective_care___hospital 
                where CAST(score as integer) <> 0 and measure_id = '"""+ measureid + """' and state =  '"""+ item[1] + """' """
                select_rows = c1.execute(sql_select_main_str)
                if len([int(i[0]) for i in select_rows]) > 1:
                    select_rows1 = c1.execute(sql_select_main_str)
                    state_sheet.cell(row = rowidx, column = 6).value = statistics.stdev(int(i[0]) for i in select_rows1)
                else:
                    #ignore rows which has only one data in list as its not possible to calculate std deviation for 1 value
                    print(measureid," : ", item[1])
        wb2.save("measures_statistics.xlsx")
        wb2.close() 
    finally:
         conn.close()
    



#used to invoke functions in sequence
def executeFunctions():
    #create directory to store dataset
    createDirectory(staging_dir_name)
    #fwtch dataset from online source and extract to directory
    fetchExtractOnlineDataset()
    #remove null data from csv files
    removeNullFromFile()
    #store data from csv file to database
    createDatabaseAndTable() 
    #create workbooks for hospital inhouse ranking and focus group
    createInHouseHospitalRankingWorkbook()
    #create Nation wise and state wise top 100 hospital ranking excel file
    createHospitalRankingWorkbook()
    #create Nation wise and state wise statistics
    measureStatistics()
    calculateStdDev()
    
executeFunctions()